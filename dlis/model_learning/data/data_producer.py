import numpy as np
import matplotlib.pyplot as plt
import math
from shapely.geometry import Polygon, Point
import alphashape.alphashape as ConcaveHull
from openpyxl import Workbook


class Vehicle:
    def __init__(self):
        # Vehicle parameters
        self.LENGTH = 0.72  # [m]
        self.LENGTH_T = 0.36  # [m]
        self.WIDTH = 0.48  # [m]
        self.BACK_TO_WHEEL = 0.36  # [m]
        self.WHEEL_LEN = 0.1  # [m]
        self.WHEEL_WIDTH = 0.02  # [m]
        self.TREAD = 0.2  # [m]
        self.WB = 0.3  # [m]
        self.ROD_LEN = 0.5  # [m]
        self.CP_OFFSET = 0.1  # [m]

    def plot_car(self, x, y, yaw, length, ax, steer=0.0, cabcolor="-r", truck_color="-k"):
        outline = np.array([[-length / 2, (length - length / 2), (length - length / 2), -length / 2, -length / 2],
                            [self.WIDTH / 2, self.WIDTH / 2, -self.WIDTH / 2, -self.WIDTH / 2, self.WIDTH / 2]])

        rr_wheel = np.array([[self.WHEEL_LEN, -self.WHEEL_LEN, -self.WHEEL_LEN, self.WHEEL_LEN, self.WHEEL_LEN],
                             [-self.WHEEL_WIDTH - self.TREAD, -self.WHEEL_WIDTH - self.TREAD,
                              self.WHEEL_WIDTH - self.TREAD, self.WHEEL_WIDTH - self.TREAD,
                              -self.WHEEL_WIDTH - self.TREAD]])
        rl_wheel = np.copy(rr_wheel)
        rl_wheel[1, :] *= -1

        rot1 = np.array([[math.cos(yaw), math.sin(yaw)],
                         [-math.sin(yaw), math.cos(yaw)]])

        outline = (outline.T.dot(rot1)).T
        rr_wheel = (rr_wheel.T.dot(rot1)).T
        rl_wheel = (rl_wheel.T.dot(rot1)).T

        outline[0, :] += x
        outline[1, :] += y
        rr_wheel[0, :] += x
        rr_wheel[1, :] += y
        rl_wheel[0, :] += x
        rl_wheel[1, :] += y

        ax.plot(np.array(outline[0, :]).flatten(),
                np.array(outline[1, :]).flatten(), truck_color)
        ax.plot(np.array(rr_wheel[0, :]).flatten(),
                np.array(rr_wheel[1, :]).flatten(), truck_color)
        ax.plot(np.array(rl_wheel[0, :]).flatten(),
                np.array(rl_wheel[1, :]).flatten(), truck_color)
        ax.plot(x, y, "*")


class RandUP:
    def __init__(self):
        # Define constants
        self.v_limit = 0.2
        self.ang_limit = 30.0
        self.time_interval = 5
        self.DT = 0.1

        # define states and inputs
        self.x = []
        self.u = []

        # Define control input limits
        self.U_min = [-self.v_limit, np.deg2rad(-self.ang_limit)]
        self.U_max = [0, np.deg2rad(self.ang_limit)]
        # self.U_max = [0, np.deg2rad(self.ang_limit)]
        # set v_max to 0 for pure backward movement

        # select random control input
        self.U_min_random = self.U_min
        self.U_max_random = self.U_max
        for i in range(self.time_interval - 1):
            self.U_min_random = np.concatenate((self.U_min_random, self.U_min), axis=None)
            self.U_max_random = np.concatenate((self.U_max_random, self.U_max), axis=None)

        # set parameters for RandUP
        self.M = 4800
        self.angle = np.array([np.random.uniform(low=-np.deg2rad(45), high=np.deg2rad(45), size=self.M)])
        self.x = np.zeros((self.M, 3))
        self.x = np.concatenate((self.x, self.angle.T), axis=1)

    @staticmethod
    def model_predictor(randup_process, vehicle, x, u):
        new_x = x.copy()

        new_x[0] += u[0] * math.cos(x[2] - x[3]) * math.cos(x[3]) * randup_process.DT
        new_x[1] += u[0] * math.cos(x[2] - x[3]) * math.sin(x[3]) * randup_process.DT
        new_x[2] += u[1] * randup_process.DT
        new_x[3] += ((u[0] / vehicle.ROD_LEN) * math.sin(x[2] - x[3]) * randup_process.DT -
                     (vehicle.CP_OFFSET * u[1] * math.cos(x[2] - x[3]) / vehicle.ROD_LEN) * randup_process.DT)
        return new_x

    @staticmethod
    def randup_trailer(vehicle, randup_process, polygon_operator):
        reachable_sets = []
        obstacles = []
        intersection_sets = []
        sampling_reach_points = []
        for i_x in range(len(randup_process.x)):
            print(i_x)
            #randup_process.x[0] = [0, 0, 0, np.deg2rad(0)]
            # x[1] = [0, 0, 0, np.deg2rad(30)]
            # x[2] = [0, 0, 0, np.deg2rad(-45)]
            # print(ind, np.rad2deg(x[ind, 3]))

            # with random control inputs but keep same during the time interval
            ys = np.zeros((randup_process.M, 4))
            # with totally random control inputs
            ys_random = np.zeros((randup_process.M, 4))
            # max velocity with bang-bang angular velocity, ysu - upper_bound w, ysl - lower_bound w, ysz - w=0
            ysu = np.zeros(4)
            ysl = np.zeros(4)
            ysz = np.zeros(4)

            us = np.random.uniform(low=randup_process.U_max, high=randup_process.U_min, size=(randup_process.M, 2))
            us_random = np.random.uniform(low=randup_process.U_max_random, high=randup_process.U_min_random,
                                          size=(randup_process.M, 2 * randup_process.time_interval))

            for t in range(randup_process.time_interval):
                if t == 0:
                    ysu = randup_process.model_predictor(randup_process, vehicle, randup_process.x[i_x],
                                                         [-randup_process.v_limit, np.deg2rad(65)])
                    ysl = randup_process.model_predictor(randup_process, vehicle, randup_process.x[i_x],
                                                         [-randup_process.v_limit, -np.deg2rad(65)])
                    ysz = randup_process.model_predictor(randup_process, vehicle, randup_process.x[i_x],
                                                         [-randup_process.v_limit, 0])
                    for i_u in range(len(us)):
                        ys[i_u, :] = randup_process.model_predictor(randup_process, vehicle, randup_process.x[i_x],
                                                                    us[i_u])
                        ys_random[i_u, :] = randup_process.model_predictor(randup_process, vehicle,
                                                                           randup_process.x[i_x],
                                                                           us_random[i_u, t * 2:t * 2 + 2])
                else:
                    ysu = randup_process.model_predictor(randup_process, vehicle, ysu,
                                                         [-randup_process.v_limit, np.deg2rad(65)])
                    ysl = randup_process.model_predictor(randup_process, vehicle, ysl,
                                                         [-randup_process.v_limit, -np.deg2rad(65)])
                    ysz = randup_process.model_predictor(randup_process, vehicle, ysz,
                                                         [-randup_process.v_limit, 0])
                    for i_u in range(len(us)):
                        ys[i_u, :] = randup_process.model_predictor(randup_process, vehicle, ys[i_u], us[i_u])
                        ys_random[i_u, :] = randup_process.model_predictor(randup_process, vehicle, ys_random[i_u],
                                                                           us_random[i_u, t * 2:t * 2 + 2])

            # insert an obstacle and calculate intersection area
            as_circle = True

            obstacle_center = np.random.uniform(low=[-5, -5], high=[5, 5], size=(1, 2))
            if as_circle:
                radius = np.linalg.norm(randup_process.x[i_x, :2]-obstacle_center)
                obstacle = Point(obstacle_center).buffer(
                    distance=np.random.normal(loc=radius, scale=0.1),
                    resolution=100)
            else:
                obstacle_vertices = (np.array([[0.5, 0], [0, 0.5], [-0.5, 0], [0, -0.5]]) *
                                     np.random.uniform(low=1, high=10))
                obstacle_vertices_random = (obstacle_vertices + obstacle_center)
                obstacle = polygon_operator.create_polygon(obstacle_vertices_random, is_obstacle=True)
            reachable_set = polygon_operator.create_polygon(ys_random[:, :2])
            intersection, intersection_area = PolygonOperator.intersection_2poly(reachable_set, obstacle)

            reachable_sets.append(reachable_set)
            obstacles.append(obstacle)
            intersection_sets.append(intersection)
            sampling_reach_points.append(ys_random)

            # plot
            """
            RandUP.plot_sets(vehicle,
                             randup_process,
                             reachable_sets,
                             obstacles,
                             intersection_sets,
                             sampling_reach_points,
                             index=i_x)
            """
        return reachable_sets, obstacles, intersection_sets, sampling_reach_points, as_circle

    @staticmethod
    def plot_sets(vehicle, randup_process, reachable_sets, obstacles, intersection_sets, sampling_reach_points, index=-1):
        if index != -1:
            ox, oy = obstacles[index].exterior.xy
            reachable_bound = np.array(reachable_sets[index].boundary.coords)
            if intersection_sets[index] is not None:
                intersection_bound = np.array(intersection_sets[index].boundary.coords)
            sampling_points = sampling_reach_points[index]

            # sub-figure configuration
            fig, axes = plt.subplots(2, 1, figsize=[8, 16])
            ax1 = axes[0]
            ax1.axis('equal')
            ax2 = axes[1]
            ax2.axis('equal')

            ax1.scatter(sampling_points[:, 0], sampling_points[:, 1], color='b')
            ax1.plot(ox, oy, color='r')

            ax1.plot(reachable_bound[:, 0], reachable_bound[:, 1], 'g')
            ax2.plot(reachable_bound[:, 0], reachable_bound[:, 1], 'g')
            # ax2.plot(intersection_bound[:, 0], intersection_bound[:, 1], 'g')

            vehicle.plot_car(randup_process.x[index, 0], randup_process.x[index, 1], randup_process.x[index, 3],
                             vehicle.LENGTH_T, ax2)

            vehicle.plot_car(randup_process.x[index, 0] + np.cos(randup_process.x[index, 3]) * vehicle.ROD_LEN +
                             np.cos(randup_process.x[index, 2]) * vehicle.CP_OFFSET,
                             randup_process.x[index, 1] + np.sin(randup_process.x[index, 3]) * vehicle.ROD_LEN +
                             np.sin(randup_process.x[index, 2]) * vehicle.CP_OFFSET,
                             randup_process.x[index, 2], vehicle.LENGTH, ax2)
            plt.show()

        else:
            assert len(reachable_sets) == len(obstacles) == len(intersection_sets) == len(sampling_reach_points), \
                'The sets do not equal one another!!'
            for i in range(len(reachable_sets)):
                ox, oy = obstacles[i].exterior.xy
                reachable_bound = np.array(reachable_sets[i].boundary.coords)
                if intersection_sets[i] is not None:
                    intersection_bound = np.array(intersection_sets[i].boundary.coords)
                sampling_points = sampling_reach_points[i]

                # sub-figure configuration
                fig, axes = plt.subplots(2, 1, figsize=[8, 16])
                ax1 = axes[0]
                ax1.axis('equal')
                ax2 = axes[1]
                ax2.axis('equal')

                ax1.scatter(sampling_points[:, 0], sampling_points[:, 1], color='b')
                ax1.fill(ox, oy, color='red')

                ax1.plot(reachable_bound[:, 0], reachable_bound[:, 1], 'g')
                ax2.plot(reachable_bound[:, 0], reachable_bound[:, 1], 'g')
                # ax2.plot(intersection_bound[:, 0], intersection_bound[:, 1], 'g')

                vehicle.plot_car(randup_process.x[i, 0], randup_process.x[i, 1], randup_process.x[i, 3],
                                 vehicle.LENGTH_T, ax2)

                vehicle.plot_car(randup_process.x[i, 0] + np.cos(randup_process.x[i, 3]) * vehicle.ROD_LEN +
                                 np.cos(randup_process.x[i, 2]) * vehicle.CP_OFFSET,
                                 randup_process.x[i, 1] + np.sin(randup_process.x[i, 3]) * vehicle.ROD_LEN +
                                 np.sin(randup_process.x[i, 2]) * vehicle.CP_OFFSET,
                                 randup_process.x[i, 2], vehicle.LENGTH, ax2)
                plt.show()


class PolygonOperator:
    def __init__(self, method_num=0, alpha=1):
        self.method_option = ['ConvexHull', 'ConcaveHull_with_Alphashape']
        self.method_num = method_num
        self.method = self.method_option[self.method_num]
        if self.method == self.method_option[1]:
            self.alpha = alpha

    def create_polygon(self, data, is_obstacle=False):
        if is_obstacle:
            polygon = Polygon(data).convex_hull
        else:
            if self.method == self.method_option[0]:
                polygon = Polygon(data).convex_hull
            elif self.method == self.method_option[1]:
                polygon = ConcaveHull(data, alpha=self.alpha)
            else:
                raise 'Please input the correct Polygon-creating method!'

        return polygon

    @staticmethod
    def intersection_2poly(reachable_set, obstacle):
        if not reachable_set.intersects(obstacle):
            inter_area = 0
            intersection = None
        else:
            intersection = reachable_set.intersection(obstacle)
            inter_area = intersection.area

        return intersection,  inter_area


def output_excel(initial_states, obstacles, intersection_sets, as_circle=False):
    wb = Workbook()
    ws_initial_states = wb.create_sheet("Initial States", 0)
    ws_obstacles = wb.create_sheet("Obstacles", 1)
    ws_intersections = wb.create_sheet("Intersection Areas", 2)
    for i in range(len(obstacles)):
        ws_initial_states.append(initial_states[i].tolist())
        if as_circle:
            obstacle_center = np.array(obstacles[i].centroid.coords)
            radius = np.array(abs(obstacles[i].bounds[3]-obstacles[i].bounds[1])/2)
            ws_obstacles.append(np.concatenate((obstacle_center, radius), axis=None).tolist())
        else:
            obstacle_coords = np.array(obstacles[i].boundary.coords).reshape((1, -1)).squeeze().tolist()
            ws_obstacles.append(obstacle_coords)
        if intersection_sets[i] is not None:
            inter_area = np.array(intersection_sets[i].area).reshape(1, 1).tolist()
            ws_intersections.append(inter_area[0])
        else:
            ws_intersections.append([0])
    wb.save("./dataset/sample.xlsx")


def main():
    randup_process = RandUP()
    vehicle = Vehicle()
    # method_num: 0 -> Convex_hull; 1 -> Concave_hull with alphashape
    polygon_operator = PolygonOperator(method_num=1, alpha=3)

    reachable_sets, obstacles, intersection_sets, sampling_reach_points, as_circle = (
        RandUP.randup_trailer(vehicle,
                              randup_process,
                              polygon_operator)
    )
    initial_states = randup_process.x
    # RandUP.plot_sets(vehicle, randup_process, reachable_sets, obstacles, intersection_sets, sampling_reach_points)
    output_excel(initial_states, obstacles, intersection_sets, as_circle)


if __name__ == '__main__':
    main()
