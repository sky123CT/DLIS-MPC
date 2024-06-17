import sys
sys.path.append('../../../')
import numpy as np
from casadi import *
import matplotlib.pyplot as plt
from shapely.geometry import Polygon, Point
import alphashape.alphashape as ConcaveHull
from openpyxl import Workbook
from mpc.functional.car import Car
from mpc.functional.dynamics import Dynamics



class PolygonOperator:
    def __init__(self, method_num=0, alpha=1.0):
        self.method_option = ['ConvexHull', 'ConcaveHull_with_Alphashape']
        self.method_num = method_num
        self.method = self.method_option[self.method_num]
        if self.method == self.method_option[1]:
            self.alpha = alpha

    def create_polygon(self, data, is_obstacle=False):
        if is_obstacle:
            polygon = Polygon(data).convex_hull
        else:
            if self.method == self.method_option[0]:
                polygon = Polygon(data).convex_hull
            elif self.method == self.method_option[1]:
                polygon = ConcaveHull(data, alpha=self.alpha)
            else:
                raise 'Please input the correct Polygon-creating method!'

        return polygon

    @staticmethod
    def intersection_2poly(reachable_set, obstacle):
        if not reachable_set.intersects(obstacle):
            inter_area = 0
            intersection = None
        else:
            intersection = reachable_set.intersection(obstacle)
            inter_area = intersection.area

        return intersection,  inter_area


class RandUP:
    def __init__(self, car, dynamics, polygon_operator, sample_num=4800, point_num=1000, scale=0.1, real_time=False):
        self.car = car
        self.dynamics = dynamics
        self.polygon_operator = polygon_operator
        self.scale = scale

        # Define constants
        self.v_limit = 0.2
        self.ang_limit = 30.0
        self.time_interval = 5

        # define states and inputs
        self.x = []
        self.u = []

        # Define control input limits
        self.U_min = [-self.v_limit, np.deg2rad(-self.ang_limit)]
        self.U_max = [0, np.deg2rad(self.ang_limit)]
        # self.U_max = [0, np.deg2rad(self.ang_limit)]
        # set v_max to 0 for pure backward movement

        # select random control input
        self.U_min_random = self.U_min
        self.U_max_random = self.U_max
        for i in range(self.time_interval - 1):
            self.U_min_random = np.concatenate((self.U_min_random, self.U_min), axis=None)
            self.U_max_random = np.concatenate((self.U_max_random, self.U_max), axis=None)

        self.point_num = point_num
        if real_time:
            self.M = 1
            self.x = np.array([[self.car.state.x,
                                self.car.state.y,
                                self.car.state.theta1,
                                self.car.state.theta2,
                                self.car.state.theta3]])
        else:
            self.M = sample_num
            relative_angle_tractor_to_trailer1 = np.array(
                [np.random.uniform(low=-np.deg2rad(45), high=np.deg2rad(45), size=self.M)])
            relative_angle_trailer1_to_trailer2 = np.array(
                [np.random.uniform(low=-np.deg2rad(30), high=np.deg2rad(30), size=self.M)])
            self.x = np.zeros((self.M, 5))
            self.x[:, 3] = self.x[:, 2] + relative_angle_tractor_to_trailer1
            self.x[:, 4] = self.x[:, 3] + relative_angle_trailer1_to_trailer2

    def onestep_model_mapping(self, u):
        new_states = np.array(
            self.dynamics.dynamics_mapping(
                x0=DM([self.car.state.x,
                       self.car.state.y,
                       self.car.state.theta1,
                       self.car.state.theta2,
                       self.car.state.theta3]),
                p=DM(u))['xf']
        ).squeeze()
        self.car.update_state(new_states[0], new_states[1], new_states[2], new_states[3], new_states[4])

    def randup_trailer(self):
        initial_states_trailer1 = []
        initial_states_trailer2 = []
        initial_states_tractor = []
        reachable_sets_trailer1 = []
        reachable_sets_trailer2 = []
        reachable_sets_trailer1_polygon = []
        reachable_sets_trailer2_polygon = []
        obstacles_for_trailer1 = []
        obstacles_for_trailer2 = []
        intersection_sets_trailer1 = []
        intersection_sets_trailer2 = []

        for i_x in range(self.M):
            print(i_x)
            reachable_set_trailer1 = []
            reachable_set_trailer2 = []

            us = np.random.uniform(low=self.U_max, high=self.U_min, size=(self.point_num, 2))
            u_random = np.random.uniform(low=self.U_min_random,
                                         high=self.U_max_random,
                                         size=(self.point_num, 2 * self.time_interval))
            for i_u in range(self.point_num):
                self.car.update_state(self.x[i_x, 0], self.x[i_x, 1], self.x[i_x, 2], self.x[i_x, 3], self.x[i_x, 4])
                if i_u == 0:
                    car_data_original, _ = self.car.get_full_data()
                    initial_states_trailer1.append(car_data_original['trailer1'])
                    initial_states_trailer2.append(car_data_original['trailer2'])
                    initial_states_tractor.append(car_data_original['tractor'])
                for t in range(self.time_interval):
                    # self.onestep_model_mapping(u=u_random[i_u, (t*2):(t*2+2)])
                    self.onestep_model_mapping(us[i_u, :])

                car_data, _ = self.car.get_full_data()
                reach_points_trailer2 = car_data['trailer2']
                reach_points_trailer1 = car_data['trailer1']
                reachable_set_trailer2.append(reach_points_trailer2[:2])
                reachable_set_trailer1.append(reach_points_trailer1[:2])

            # reachable sampling points
            reachable_set_trailer1 = np.array(reachable_set_trailer1)
            reachable_set_trailer2 = np.array(reachable_set_trailer2)

            # obstacle polygon
            obstacle_center = np.random.uniform(low=[-5, -5], high=[5, 5], size=(1, 2))
            radius_for_trailer1 = np.linalg.norm(initial_states_trailer1[i_x][:2] - obstacle_center)
            radius_for_trailer2 = np.linalg.norm(initial_states_trailer2[i_x][:2] - obstacle_center)
            random_radius1 = np.random.normal(loc=radius_for_trailer1, scale=self.scale)
            random_radius2 = (radius_for_trailer2 +
                              ((random_radius1-radius_for_trailer1) / radius_for_trailer1) * radius_for_trailer2)
            obstacle_for_trailer1 = Point(obstacle_center).buffer(
                distance=random_radius1, resolution=100)
            obstacle_for_trailer2 = Point(obstacle_center).buffer(
                distance=random_radius2, resolution=100)

            # reachable polygon and intersection polygon
            reachable_set_trailer1_polygon = self.polygon_operator.create_polygon(reachable_set_trailer1)
            reachable_set_trailer2_polygon = self.polygon_operator.create_polygon(reachable_set_trailer2)
            intersection_trailer1, intersection_area_trailer1 = self.polygon_operator.intersection_2poly(
                reachable_set_trailer1_polygon, obstacle_for_trailer1)
            intersection_trailer2, intersection_area_trailer2 = self.polygon_operator.intersection_2poly(
                reachable_set_trailer2_polygon, obstacle_for_trailer2)

            # append into lists
            reachable_sets_trailer1.append(reachable_set_trailer1)
            reachable_sets_trailer2.append(reachable_set_trailer2)
            reachable_sets_trailer1_polygon.append(reachable_set_trailer1_polygon)
            reachable_sets_trailer2_polygon.append(reachable_set_trailer2_polygon)
            obstacles_for_trailer1.append(obstacle_for_trailer1)
            obstacles_for_trailer2.append(obstacle_for_trailer2)
            intersection_sets_trailer1.append(intersection_trailer1)
            intersection_sets_trailer2.append(intersection_trailer2)


            """self.plot_sets(sampling_points1=reachable_set_trailer1,
                           sampling_points2=reachable_set_trailer2,
                           reachable_set1=reachable_set_trailer1_polygon,
                           reachable_set2=reachable_set_trailer2_polygon,
                           obstacle1=obstacle_for_trailer1,
                           obstacle2=obstacle_for_trailer2,
                           current_original_state=self.x[i_x, :])
"""

        return (initial_states_trailer1,
                initial_states_trailer2,
                initial_states_tractor,
                reachable_sets_trailer1,
                reachable_sets_trailer2,
                reachable_sets_trailer1_polygon,
                reachable_sets_trailer2_polygon,
                obstacles_for_trailer1,
                obstacles_for_trailer2,
                intersection_sets_trailer1,
                intersection_sets_trailer2)

    def plot_sets(self,
                  sampling_points1,
                  sampling_points2,
                  reachable_set1,
                  reachable_set2,
                  obstacle1,
                  obstacle2,
                  current_original_state):
        sampling_points1 = np.array(sampling_points1)
        sampling_points2 = np.array(sampling_points2)
        ox1, oy1 = obstacle1.exterior.xy
        ox2, oy2 = obstacle2.exterior.xy
        reachable_bound1 = np.array(reachable_set1.boundary.coords)
        reachable_bound2 = np.array(reachable_set2.boundary.coords)
        self.car.update_state(
            current_original_state[0],
            current_original_state[1],
            current_original_state[2],
            current_original_state[3],
            current_original_state[4]
        )

        plt.clf()

        plt.subplot(321)
        plt.axis("equal")
        plt.plot(ox1, oy1, "-r", label='obstacle_for_trailer1')
        plt.plot(reachable_bound1[:, 0], reachable_bound1[:, 1], '-g', label='reachable_polygon1')
        plt.legend()

        plt.subplot(322)
        plt.axis("equal")
        plt.plot(ox2, oy2, "-r", label='obstacle_for_trailer2')
        plt.plot(reachable_bound2[:, 0], reachable_bound2[:, 1], '-g', label='reachable_polygon2')
        plt.legend()

        plt.subplot(323)
        plt.axis('equal')
        plt.plot(reachable_bound1[:, 0], reachable_bound1[:, 1], '-g', label='reachable_polygon1')
        self.car.plot_full_car()
        plt.legend()

        plt.subplot(324)
        plt.axis('equal')
        plt.plot(reachable_bound2[:, 0], reachable_bound2[:, 1], '-g', label='reachable_polygon2')
        self.car.plot_full_car()
        plt.legend()

        plt.subplot(325)
        plt.axis('equal')
        plt.scatter(sampling_points1[:, 0], sampling_points1[:, 1], color='b', label='sampling_points1')
        plt.plot(reachable_bound1[:, 0], reachable_bound1[:, 1], '-g', label='reachable_polygon1')
        plt.legend()

        plt.subplot(326)
        plt.axis('equal')
        plt.scatter(sampling_points2[:, 0], sampling_points2[:, 1], color='b', label='sampling_points2')
        plt.plot(reachable_bound2[:, 0], reachable_bound2[:, 1], '-g', label='reachable_polygon2')
        plt.legend()

        plt.show()
        plt.pause(0.1)

    def output_as_excel(self,
                        initial_states_trailer1,
                        initial_states_trailer2,
                        initial_states_tractor,
                        obstacles1,
                        obstacles2,
                        intersection_sets1,
                        intersection_sets2,):
        wb = Workbook()
        ws_initial_states = wb.create_sheet("Initial_States", 0)
        ws_obstacles = wb.create_sheet("Obstacles", 1)
        ws_intersections = wb.create_sheet("Intersection_Areas", 2)
        assert len(obstacles1) == len(obstacles2) == len(intersection_sets1) == len(intersection_sets2) == self.M, \
            "sample number not equal!"
        for i in range(self.M):
            ws_initial_states.append(np.concatenate(
                (initial_states_trailer1[i], initial_states_trailer2[i], initial_states_tractor[i])
            ).tolist())
            obstacle_center = np.array(obstacles1[i].centroid.coords)
            radius1 = np.array(abs(obstacles1[i].bounds[3] - obstacles1[i].bounds[1]) / 2)
            radius2 = np.array(abs(obstacles2[i].bounds[3] - obstacles2[i].bounds[1]) / 2)
            ws_obstacles.append(np.concatenate((obstacle_center, radius1, radius2), axis=None).tolist())

            if intersection_sets1[i] is not None:
                inter_area1 = np.array(intersection_sets1[i].area).reshape(1, 1)
            else:
                inter_area1 = np.zeros((1, 1))
            if intersection_sets2[i] is not None:
                inter_area2 = np.array(intersection_sets2[i].area).reshape(1, 1)
            else:
                inter_area2 = np.zeros((1, 1))
            ws_intersections.append(np.concatenate((inter_area1, inter_area2), axis=None).tolist())

        wb.save("./dataset/sample_2trailers_" +
                str(self.M) + "samples_with_tractor_" + str(self.point_num) + "points_" +
                "scale_" + str(self.scale) + ".xlsx")


def main():
    argv = sys.argv
    sample_num = 4800 #int(argv[1])
    point_num = 1000 #int(argv[2])
    scale = 0.05 #int(argv[3])

    car = Car(trailer_num=3)
    dynamics = Dynamics(car=car, nx=5, nu=2, dt=0.1)
    polygon_operator = PolygonOperator(method_num=1, alpha=0.05)
    randup_process = RandUP(car=car,
                            dynamics=dynamics,
                            polygon_operator=polygon_operator,
                            sample_num=sample_num,
                            point_num=point_num,
                            scale=scale)

    (initial_states_trailer1,
     initial_states_trailer2,
     initial_states_tractor,
     reachable_sets_trailer1,
     reachable_sets_trailer2,
     reachable_sets_trailer1_polygon,
     reachable_sets_trailer2_polygon,
     obstacles_for_trailer1,
     obstacles_for_trailer2,
     intersection_sets_trailer1,
     intersection_sets_trailer2) = randup_process.randup_trailer()

    randup_process.output_as_excel(initial_states_trailer1=initial_states_trailer1,
                                   initial_states_trailer2=initial_states_trailer2,
                                   initial_states_tractor=initial_states_tractor,
                                   obstacles1=obstacles_for_trailer1,
                                   obstacles2=obstacles_for_trailer2,
                                   intersection_sets1=intersection_sets_trailer1,
                                   intersection_sets2=intersection_sets_trailer2)


if __name__ == '__main__':
    main()



